# gestion/views.py
from datetime import date, timedelta
import io

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.utils import timezone

from .models import ReleveCentreLivre, Livre, Centre
from .forms import ReleveCentreLivreForm, LivreForm, CentreForm

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from xhtml2pdf import pisa


from django.core.exceptions import PermissionDenied
from functools import wraps

from django.shortcuts import get_object_or_404
from django.core.exceptions import PermissionDenied


def admin_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        user = request.user
        if not user.is_authenticated:
            raise PermissionDenied
        # admin = rôle ADMIN ou superuser
        if getattr(user, "is_superuser", False):
            return view_func(request, *args, **kwargs)
        if hasattr(user, "is_admin") and user.is_admin():
            return view_func(request, *args, **kwargs)
        raise PermissionDenied
    return _wrapped


def _week_range(year: int, week: int):
    """Retourne (date_debut, date_fin) pour une semaine ISO (lundi → dimanche)."""
    # fromisocalendar(year, week, weekday) : weekday = 1 (lundi)
    start = date.fromisocalendar(year, week, 1)
    end = start + timedelta(days=6)
    return start, end


def _month_range(year: int, month: int):
    """Retourne (date_debut, date_fin) pour un mois donné."""
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


@login_required
def rapport_hebdomadaire(request):
    today = timezone.now().date()
    iso = today.isocalendar()  # namedtuple (year, week, weekday)

    try:
        year = int(request.GET.get("year", iso.year))
    except ValueError:
        year = iso.year

    try:
        week = int(request.GET.get("week", iso.week))
    except ValueError:
        week = iso.week

    date_debut, date_fin = _week_range(year, week)

    # uniquement relevés de type "semaine"
    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    current_year = today.year
    year_choices = list(range(current_year - 5, current_year + 2))
    week_choices = [(i, f"Semaine {i}") for i in range(1, 54)]

    label = f"Semaine {week} ({date_debut} → {date_fin})"

    context = {
        "year": year,
        "week": week,
        "week_label": label,
        "year_choices": year_choices,
        "week_choices": week_choices,
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }
    return render(request, "gestion/rapport_hebdomadaire.html", context)


@login_required
def rapport_mensuel(request):
    # année / mois depuis les paramètres GET ou date du jour
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year

    try:
        month = int(request.GET.get("month", today.month))
    except ValueError:
        month = today.month

    # bornes du mois
    date_debut, date_fin = _month_range(year, month)

    # queryset de base : relevés du mois, type "mois"
    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user

    # si utilisateur centre, on limite aux relevés & centres de ce centre
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    # nos 4 livres (plus tard tu pourras en ajouter d'autres si besoin)
    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    # choix année & mois pour le formulaire de filtre
    current_year = today.year
    year_choices = list(range(current_year - 5, current_year + 2))
    month_choices = [
        (1, "Janvier"),
        (2, "Février"),
        (3, "Mars"),
        (4, "Avril"),
        (5, "Mai"),
        (6, "Juin"),
        (7, "Juillet"),
        (8, "Août"),
        (9, "Septembre"),
        (10, "Octobre"),
        (11, "Novembre"),
        (12, "Décembre"),
    ]
    month_name = dict(month_choices).get(month, "")

    context = {
        "year": year,
        "month": month,
        "month_name": month_name,
        "year_choices": year_choices,
        "month_choices": month_choices,
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }
    return render(request, "gestion/rapport_mensuel.html", context)


def _quarter_range(year: int, quarter: int):
    """Retourne (date_debut, date_fin) pour un trimestre donné (1 à 4)."""
    from math import ceil
    if quarter < 1:
        quarter = 1
    if quarter > 4:
        quarter = 4
    start_month = (quarter - 1) * 3 + 1  # 1, 4, 7, 10
    end_month = start_month + 2          # 3, 6, 9, 12
    start, _ = _month_range(year, start_month)
    _, end = _month_range(year, end_month)
    return start, end


@login_required
def rapport_trimestriel(request):
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year

    try:
        quarter = int(request.GET.get("quarter", 1))
    except ValueError:
        quarter = 1

    date_debut, date_fin = _quarter_range(year, quarter)

    # Tous les relevés dont la date_fin est dans le trimestre
    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    current_year = today.year
    year_choices = list(range(current_year - 5, current_year + 2))
    quarter_choices = [
        (1, "1er trimestre (Jan–Mar)"),
        (2, "2e trimestre (Avr–Jun)"),
        (3, "3e trimestre (Jul–Sep)"),
        (4, "4e trimestre (Oct–Déc)"),
    ]
    quarter_label = dict(quarter_choices).get(quarter, "")

    context = {
        "year": year,
        "quarter": quarter,
        "quarter_label": quarter_label,
        "year_choices": year_choices,
        "quarter_choices": quarter_choices,
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }
    return render(request, "gestion/rapport_trimestriel.html", context)

@login_required
def rapport_annuel(request):
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year

    # bornes de l'année
    date_debut = date(year, 1, 1)
    date_fin = date(year, 12, 31)

    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    current_year = today.year
    year_choices = list(range(current_year - 5, current_year + 2))

    context = {
        "year": year,
        "year_choices": year_choices,
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }
    return render(request, "gestion/rapport_annuel.html", context)


@login_required
def rapport_global(request):
    # Tous les relevés (toutes périodes confondues)
    releves = ReleveCentreLivre.objects.all()

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    context = {
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }
    return render(request, "gestion/rapport_global.html", context)


def _export_via_act_excel_from_releves(releves, centres, sheet_title, filename):
    """
    Construit un fichier Excel au format VIAT/ACT (comme ton modèle)
    à partir d'un queryset de relevés + liste de centres.
    """
    # On cible spécifiquement les livres "Viatique" et "Activités"
    viatique = (
        Livre.objects.filter(nom__iexact="Viatique").first()
        or Livre.objects.filter(code__icontains="VIAT").first()
    )
    activite = (
        Livre.objects.filter(nom__iexact="Activités").first()
        or Livre.objects.filter(code__icontains="ACT").first()
    )

    # Fonctions utilitaires
    def agg_livre_global(livre):
        if not livre:
            return {"q_recue": 0, "q_vendue": 0, "reste": 0}
        agg = releves.filter(livre=livre).aggregate(
            q_recue=Sum("quantite_recue"),
            q_vendue=Sum("quantite_vendue"),
            reste=Sum("quantite_reste"),
        )
        return {
            "q_recue": agg["q_recue"] or 0,
            "q_vendue": agg["q_vendue"] or 0,
            "reste": agg["reste"] or 0,
        }

    def agg_livre_centre(centre, livre):
        if not livre:
            return {"q_recue": 0, "q_vendue": 0, "reste": 0, "montant": 0, "dep": 0}
        agg = releves.filter(centre=centre, livre=livre).aggregate(
            q_recue=Sum("quantite_recue"),
            q_vendue=Sum("quantite_vendue"),
            reste=Sum("quantite_reste"),
            montant=Sum("montant_ventes"),
            dep=Sum("depenses"),
        )
        return {
            "q_recue": agg["q_recue"] or 0,
            "q_vendue": agg["q_vendue"] or 0,
            "reste": agg["reste"] or 0,
            "montant": agg["montant"] or 0,
            "dep": agg["dep"] or 0,
        }

    # Totaux globaux pour les 2 lignes du haut
    g_viat = agg_livre_global(viatique)
    g_act = agg_livre_global(activite)

    # --- Création du fichier Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    fill_viat_top = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")   # vert
    fill_act_top = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")    # jaune
    fill_blue = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    fill_green = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_white_bold = Font(color="FFFFFF", bold=True)

    # Lignes 2–3 : résumé global par livre
    # B2:E2 = VIATIQUE
    ws["B2"].value = "VIATIQUE"
    ws["B2"].font = header_font
    ws["B2"].alignment = center
    ws["B2"].fill = fill_viat_top

    ws["C2"].value = g_viat["q_recue"]
    ws["D2"].value = g_viat["q_vendue"]
    ws["E2"].value = g_viat["reste"]
    for col in ("C", "D", "E"):
        cell = ws[f"{col}2"]
        cell.alignment = right
        cell.fill = fill_viat_top

    # B3:E3 = ACTIVITE
    ws["B3"].value = "ACTIVITE"
    ws["B3"].font = header_font
    ws["B3"].alignment = center
    ws["B3"].fill = fill_act_top

    ws["C3"].value = g_act["q_recue"]
    ws["D3"].value = g_act["q_vendue"]
    ws["E3"].value = g_act["reste"]
    for col in ("C", "D", "E"):
        cell = ws[f"{col}3"]
        cell.alignment = right
        cell.fill = fill_act_top

    # Ligne d’en-tête principale (ligne 5)
    header_row = 5
    headers = [
        "CENTRE",     # A
        "VIAT",       # B (quantité reçue viat)
        "ACT",        # C (quantité reçue act)
        "VEN VIAT",   # D
        "PU",         # E (prix unitaire viat)
        "MTANT",      # F (montant viat)
        "VENT ACT",   # G
        "PU",         # H (prix unitaire act)
        "MTANT",      # I (montant act)
        "RES VIA",    # J
        "RES ACT",    # K
        "DEPENSES",   # L
    ]
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.font = header_font
        cell.alignment = center
        if label in ("VIAT", "VEN VIAT", "PU", "MTANT", "RES VIA"):
            cell.fill = fill_blue
        elif label in ("ACT", "VENT ACT", "PU", "MTANT", "RES ACT"):
            cell.fill = fill_green
        elif label == "DEPENSES":
            cell.fill = fill_yellow

    # Lignes de données (centres)
    first_data_row = header_row + 1
    row_idx = first_data_row

    # On utilisera les prix unitaires par défaut des livres
    pu_viat = float(viatique.prix_unitaire_defaut) if viatique else 0
    pu_act = float(activite.prix_unitaire_defaut) if activite else 0

    for centre in centres:
        s_viat = agg_livre_centre(centre, viatique)
        s_act = agg_livre_centre(centre, activite)

        ws.cell(row=row_idx, column=1, value=centre.nom)  # CENTRE

        ws.cell(row=row_idx, column=2, value=s_viat["q_recue"])   # VIAT
        ws.cell(row=row_idx, column=3, value=s_act["q_recue"])    # ACT

        ws.cell(row=row_idx, column=4, value=s_viat["q_vendue"])  # VEN VIAT
        ws.cell(row=row_idx, column=5, value=pu_viat)             # PU (viat)
        ws.cell(row=row_idx, column=6, value=float(s_viat["montant"]))  # MTANT (viat)

        ws.cell(row=row_idx, column=7, value=s_act["q_vendue"])   # VENT ACT
        ws.cell(row=row_idx, column=8, value=pu_act)              # PU (act)
        ws.cell(row=row_idx, column=9, value=float(s_act["montant"]))   # MTANT (act)

        ws.cell(row=row_idx, column=10, value=s_viat["reste"])    # RES VIA
        ws.cell(row=row_idx, column=11, value=s_act["reste"])     # RES ACT

        ws.cell(row=row_idx, column=12, value=float(s_viat["dep"] + s_act["dep"]))  # DEPENSES

        for col in range(2, 13):
            ws.cell(row=row_idx, column=col).alignment = right

        row_idx += 1

    last_data_row = row_idx - 1

    # Ligne TOTAL
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = font_white_bold
    ws.cell(row=total_row, column=1).fill = fill_red
    ws.cell(row=total_row, column=1).alignment = center

    for col in range(2, 13):
        letter = get_column_letter(col)
        cell = ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})",
        )
        cell.font = font_white_bold
        cell.fill = fill_red
        cell.alignment = right

    # Ligne SOMME
    somme_row = total_row + 2
    ws.cell(row=somme_row, column=5, value="SOMME")
    ws.cell(row=somme_row, column=5).font = font_white_bold
    ws.cell(row=somme_row, column=5).fill = fill_red
    ws.cell(row=somme_row, column=5).alignment = center

    ws.cell(
        row=somme_row,
        column=6,
        value=f"=SUM(F{first_data_row}:F{last_data_row})+SUM(I{first_data_row}:I{last_data_row})",
    )
    ws.cell(row=somme_row, column=6).font = font_white_bold
    ws.cell(row=somme_row, column=6).fill = fill_red
    ws.cell(row=somme_row, column=6).alignment = right

    # Largeurs colonnes
    widths = {
        1: 18, 2: 8, 3: 8, 4: 10, 5: 8, 6: 12,
        7: 10, 8: 8, 9: 12, 10: 10, 11: 10, 12: 12,
    }
    for col_idx, width in widths.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

    # Réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_rapport_mensuel_excel(request):
    """Export mensuel VIAT/ACT"""
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year
    try:
        month = int(request.GET.get("month", today.month))
    except ValueError:
        month = today.month

    date_debut, date_fin = _month_range(year, month)

    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    sheet_title = f"{year}-{month:02d}"
    filename = f"rapport_mensuel_via_act_{year}_{month:02d}.xlsx"
    return _export_via_act_excel_from_releves(releves, centres, sheet_title, filename)


@login_required
def export_rapport_trimestriel_excel(request):
    """Export trimestriel VIAT/ACT"""
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year
    try:
        quarter = int(request.GET.get("quarter", 1))
    except ValueError:
        quarter = 1

    date_debut, date_fin = _quarter_range(year, quarter)

    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    sheet_title = f"T{quarter}-{year}"
    filename = f"rapport_trimestriel_via_act_{year}_T{quarter}.xlsx"
    return _export_via_act_excel_from_releves(releves, centres, sheet_title, filename)


@login_required
def export_rapport_annuel_excel(request):
    """Export annuel VIAT/ACT"""
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year

    date_debut = date(year, 1, 1)
    date_fin = date(year, 12, 31)

    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    sheet_title = str(year)
    filename = f"rapport_annuel_via_act_{year}.xlsx"
    return _export_via_act_excel_from_releves(releves, centres, sheet_title, filename)


@login_required
def export_rapport_global_excel(request):
    """Export global VIAT/ACT – tous les relevés"""
    releves = ReleveCentreLivre.objects.all()

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    sheet_title = "Global"
    filename = "rapport_global_via_act.xlsx"
    return _export_via_act_excel_from_releves(releves, centres, sheet_title, filename)


@login_required
def export_rapport_mensuel_pdf(request):
    """Génère un PDF du rapport mensuel (même données que la page HTML)."""
    today = timezone.now().date()

    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year

    try:
        month = int(request.GET.get("month", today.month))
    except ValueError:
        month = today.month

    # bornes du mois
    date_debut, date_fin = _month_range(year, month)

    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    month_choices = [
        (1, "Janvier"),
        (2, "Février"),
        (3, "Mars"),
        (4, "Avril"),
        (5, "Mai"),
        (6, "Juin"),
        (7, "Juillet"),
        (8, "Août"),
        (9, "Septembre"),
        (10, "Octobre"),
        (11, "Novembre"),
        (12, "Décembre"),
    ]
    month_name = dict(month_choices).get(month, "")

    context = {
        "year": year,
        "month": month,
        "month_name": month_name,
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }

    html = render_to_string("gestion/pdf_rapport_mensuel.html", context)
    response = HttpResponse(content_type="application/pdf")
    filename = f"rapport_mensuel_{year}_{month:02d}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    pisa.CreatePDF(html, dest=response)
    return response


@login_required
def export_rapport_trimestriel_pdf(request):
    """PDF du rapport trimestriel."""
    today = timezone.now().date()

    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year

    try:
        quarter = int(request.GET.get("quarter", 1))
    except ValueError:
        quarter = 1

    date_debut, date_fin = _quarter_range(year, quarter)

    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    quarter_choices = [
        (1, "1er trimestre (Jan–Mar)"),
        (2, "2e trimestre (Avr–Jun)"),
        (3, "3e trimestre (Jul–Sep)"),
        (4, "4e trimestre (Oct–Déc)"),
    ]
    quarter_label = dict(quarter_choices).get(quarter, "")

    context = {
        "year": year,
        "quarter": quarter,
        "quarter_label": quarter_label,
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }

    html = render_to_string("gestion/pdf_rapport_trimestriel.html", context)
    response = HttpResponse(content_type="application/pdf")
    filename = f"rapport_trimestriel_{year}_T{quarter}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    pisa.CreatePDF(html, dest=response)
    return response



@login_required
def export_rapport_annuel_pdf(request):
    """PDF du rapport annuel."""
    today = timezone.now().date()

    try:
        year = int(request.GET.get("year", today.year))
    except ValueError:
        year = today.year

    date_debut = date(year, 1, 1)
    date_fin = date(year, 12, 31)

    releves = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        releves = releves.filter(centre=user.centre)
        centres = Centre.objects.filter(pk=user.centre_id)
    else:
        centres = Centre.objects.all().order_by("nom")

    livres = list(Livre.objects.all().order_by("nom"))

    rows = []
    for centre in centres:
        row = {"centre": centre, "livres": []}
        for livre in livres:
            agg = releves.filter(centre=centre, livre=livre).aggregate(
                q_recue=Sum("quantite_recue"),
                q_vendue=Sum("quantite_vendue"),
                montant=Sum("montant_ventes"),
                reste=Sum("quantite_reste"),
            )
            row["livres"].append(
                {
                    "livre": livre,
                    "q_recue": agg["q_recue"] or 0,
                    "q_vendue": agg["q_vendue"] or 0,
                    "reste": agg["reste"] or 0,
                    "montant": agg["montant"] or 0,
                }
            )
        rows.append(row)

    context = {
        "year": year,
        "livres": livres,
        "rows": rows,
        "colspan": 1 + len(livres) * 4,
    }

    html = render_to_string("gestion/pdf_rapport_annuel.html", context)
    response = HttpResponse(content_type="application/pdf")
    filename = f"rapport_annuel_{year}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    pisa.CreatePDF(html, dest=response)
    return response


@login_required
@admin_required
def livre_list(request):
    livres = Livre.objects.all().order_by("nom")
    context = {
        "livres": livres,
    }
    return render(request, "gestion/livre_list.html", context)


@login_required
@admin_required
def livre_create(request):
    if request.method == "POST":
        form = LivreForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("livre_list")
    else:
        form = LivreForm()

    return render(request, "gestion/livre_form.html", {"form": form, "mode": "create"})


@login_required
@admin_required
def livre_update(request, pk):
    livre = Livre.objects.get(pk=pk)
    if request.method == "POST":
        form = LivreForm(request.POST, instance=livre)
        if form.is_valid():
            form.save()
            return redirect("livre_list")
    else:
        form = LivreForm(instance=livre)

    return render(
        request,
        "gestion/livre_form.html",
        {"form": form, "mode": "update", "livre": livre},
    )


@login_required
@admin_required
def livre_delete(request, pk):
    livre = Livre.objects.get(pk=pk)
    if request.method == "POST":
        livre.delete()
        return redirect("livre_list")
    return render(
        request,
        "gestion/livre_confirm_delete.html",
        {"livre": livre},
    )


@login_required
@admin_required
def centre_list(request):
    centres = Centre.objects.all().order_by("nom")
    return render(request, "gestion/centre_list.html", {"centres": centres})


@login_required
@admin_required
def centre_create(request):
    if request.method == "POST":
        form = CentreForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("centre_list")
    else:
        form = CentreForm()
    return render(request, "gestion/centre_form.html", {"form": form, "mode": "create"})


@login_required
@admin_required
def centre_update(request, pk):
    centre = Centre.objects.get(pk=pk)
    if request.method == "POST":
        form = CentreForm(request.POST, instance=centre)
        if form.is_valid():
            form.save()
            return redirect("centre_list")
    else:
        form = CentreForm(instance=centre)
    return render(
        request,
        "gestion/centre_form.html",
        {"form": form, "mode": "update", "centre": centre},
    )


@login_required
@admin_required
def centre_delete(request, pk):
    centre = Centre.objects.get(pk=pk)
    if request.method == "POST":
        centre.delete()
        return redirect("centre_list")
    return render(
        request,
        "gestion/centre_confirm_delete.html",
        {"centre": centre},
    )


















@login_required
def dashboard(request):
    today = timezone.now().date()
    year = today.year
    month = today.month

    date_debut, date_fin = _month_range(year, month)

    # Tous les relevés dont la date de fin est dans le mois courant
    qs = ReleveCentreLivre.objects.filter(
        date_fin__gte=date_debut,
        date_fin__lte=date_fin,
    )

    user = request.user
    # Si c'est un utilisateur "centre", il ne voit que ses relevés
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        qs = qs.filter(centre=user.centre)

    # Total des ventes du mois
    total_mois = qs.aggregate(total=Sum("montant_ventes"))["total"] or 0

    # Meilleur centre (par montant de ventes)
    ventes_par_centre = (
        qs.values("centre__nom")
        .annotate(total=Sum("montant_ventes"))
        .order_by("-total")
    )
    meilleur_centre = ventes_par_centre[0] if ventes_par_centre else None

    # Livres : quantité vendue & montant
    ventes_par_livre = (
        qs.values("livre__nom")
        .annotate(
            quantite_vendue=Sum("quantite_vendue"),
            montant_total=Sum("montant_ventes"),
        )
        .order_by("-quantite_vendue")
    )
    livre_top = ventes_par_livre[0] if ventes_par_livre else None

    month_choices = [
        (1, "Janvier"),
        (2, "Février"),
        (3, "Mars"),
        (4, "Avril"),
        (5, "Mai"),
        (6, "Juin"),
        (7, "Juillet"),
        (8, "Août"),
        (9, "Septembre"),
        (10, "Octobre"),
        (11, "Novembre"),
        (12, "Décembre"),
    ]
    month_name = dict(month_choices).get(month, "")

    context = {
        "year": year,
        "month": month,
        "month_name": month_name,
        "total_mois": total_mois,
        "meilleur_centre": meilleur_centre,
        "livre_top": livre_top,
        "ventes_livres": ventes_par_livre,
    }
    return render(request, "gestion/dashboard.html", context)



@login_required
def releve_list(request):
    qs = (
        ReleveCentreLivre.objects
        .select_related("centre", "livre")
        .order_by("-date_fin", "centre__nom")
    )

    user = request.user
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        qs = qs.filter(centre=user.centre)

    return render(request, "gestion/releve_list.html", {"releves": qs})


@login_required
def releve_create(request):
    user = request.user

    if request.method == "POST":
        form = ReleveCentreLivreForm(request.POST, user=user)
        if form.is_valid():
            releve = form.save(commit=False)

            # Pour un utilisateur "centre", on force le centre depuis le compte
            if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
                releve.centre = user.centre

            releve.save()
            return redirect("releve_list")
    else:
        form = ReleveCentreLivreForm(user=user)

    return render(request, "gestion/releve_form.html", {"form": form})


@login_required
def releve_update(request, pk):
    user = request.user
    releve = get_object_or_404(ReleveCentreLivre, pk=pk)

    # Un utilisateur "centre" ne peut modifier que ses propres relevés
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        if releve.centre_id != user.centre_id:
            raise PermissionDenied

    if request.method == "POST":
        form = ReleveCentreLivreForm(request.POST, instance=releve, user=user)
        if form.is_valid():
            obj = form.save(commit=False)
            # Sécurité : si c'est un centre, on force le centre
            if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
                obj.centre = user.centre
            obj.save()
            return redirect("releve_list")
    else:
        form = ReleveCentreLivreForm(instance=releve, user=user)

    return render(
        request,
        "gestion/releve_form.html",
        {"form": form, "mode": "update", "releve": releve},
    )


@login_required
def releve_delete(request, pk):
    user = request.user
    releve = get_object_or_404(ReleveCentreLivre, pk=pk)

    # Un utilisateur "centre" ne peut supprimer que ses propres relevés
    if hasattr(user, "is_centre") and user.is_centre() and user.centre_id:
        if releve.centre_id != user.centre_id:
            raise PermissionDenied

    if request.method == "POST":
        releve.delete()
        return redirect("releve_list")

    return render(
        request,
        "gestion/releve_confirm_delete.html",
        {"releve": releve},
        )
